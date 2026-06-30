#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generador PAQUETE DE SEPARACION - VIVA PRO (LA HAUS CONSTRUCTORA S.A.C.)
Plantillas docx -> PDF + simulacion (xlsx->PDF). La ficha inserta el plano por tipologia.
Uso: python3 generar_paquete.py cliente.json
"""
import sys, os, json, zipfile, shutil, subprocess, datetime, io
from num2words import num2words
import openpyxl
from openpyxl.worksheet.properties import PageSetupProperties
from PIL import Image

BASE = os.path.dirname(os.path.abspath(__file__))
TPL  = BASE
# Carpeta de planos: por defecto relativa a la raiz del proyecto; se puede pasar planos_dir en el cfg
DEFAULT_PLANOS = BASE
MESES = ["enero","febrero","marzo","abril","mayo","junio","julio",
         "agosto","septiembre","octubre","noviembre","diciembre"]

def m2(x):  return f"{round(x):,.2f}"
def m0(x):  return f"{round(x):,.0f}"
def pct(x): return f"{x:.2f} %"
def palabras_soles(x): return num2words(int(round(x)), lang="es").upper() + " CON 00/100 SOLES"
def fecha_larga(d): return f"{d.day} de {MESES[d.month-1]} de {d.year}"
def mes_anio(d):    return f"{MESES[d.month-1]} {d.year}"

def fit_canvas(src, cw=1536, ch=2752):
    """Devuelve bytes JPEG del plano escalado y centrado en lienzo blanco cw x ch (sin deformar)."""
    im = Image.open(src).convert("RGB")
    w, h = im.size
    s = min(cw/w, ch/h)
    nw, nh = int(w*s), int(h*s)
    im2 = im.resize((nw, nh), Image.LANCZOS)
    canvas = Image.new("RGB", (cw, ch), (255,255,255))
    canvas.paste(im2, ((cw-nw)//2, (ch-nh)//2))
    buf = io.BytesIO(); canvas.save(buf, "JPEG", quality=90); return buf.getvalue()

def render_docx(tpl_path, out_path, repl, media_swap=None):
    shutil.copyfile(tpl_path, out_path)
    tmp = out_path + ".tmp"
    with zipfile.ZipFile(out_path,"r") as zin, zipfile.ZipFile(tmp,"w",zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "word/document.xml":
                xml = data.decode("utf-8")
                for old,new in sorted(repl.items(), key=lambda kv:-len(kv[0])):
                    xml = xml.replace(old,new)
                data = xml.encode("utf-8")
            elif media_swap and item.filename in media_swap:
                data = media_swap[item.filename]
            zout.writestr(item, data)
    os.replace(tmp, out_path)

import shutil as _sh
def _find_soffice():
    for c in ("libreoffice","soffice","soffice.exe"):
        p=_sh.which(c)
        if p: return p
    for p in (r"C:\\Program Files\\LibreOffice\\program\\soffice.exe",
              r"C:\\Program Files (x86)\\LibreOffice\\program\\soffice.exe",
              "/Applications/LibreOffice.app/Contents/MacOS/soffice"):
        if os.path.exists(p): return p
    raise RuntimeError("No se encontró LibreOffice. Instálalo desde https://www.libreoffice.org/download/ y vuelve a intentar.")
def to_pdf(path, out_dir):
    so=_find_soffice()
    env=dict(os.environ); env.setdefault("HOME", os.path.expanduser("~") or "/tmp")
    subprocess.run([so,"--headless","--convert-to","pdf","--outdir",out_dir,path],
                   check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, env=env)

def construir(cfg, dep):
    P=float(cfg["precio_soles"]); S=float(cfg.get("separacion_soles",3500)); sep_usd=float(cfg.get("separacion_usd",1000))
    p20=P*0.20; p20_sep=p20-S; inicial40=P*0.40; armada=P*0.15; saldo60=P-inicial40; half50=P*0.50; p10=P*0.10
    d=datetime.date.fromisoformat(cfg["fecha"]); yyyy=d.year
    num=dep["codigo"]; piso=dep["piso"]; tip=dep["tipologia"]; area=dep["area_m2"]
    dorm=dep["dormitorios_txt"]; banos=dep["banos_txt"]; un=cfg.get("unidad_n","")
    nd={"un (01) dormitorio":1,"dos (02) dormitorios":2,"tres (03) dormitorios":3}.get(dorm,3)
    nd_txt=f"{nd} dormitorio"+("s" if nd!=1 else "")
    sexo=cfg.get("sexo","F").upper(); ape=cfg.get("apellido",""); nombre=cfg["nombre"].upper(); ec=cfg.get("estado_civil","")
    if sexo=="F": trato="Señora:"; estimado=f"Estimada Sra. {ape}"
    else:         trato="Señor:";  estimado=f"Estimado Sr. {ape}"
    R={
        "YADIRA LISSET CABALLERO NOEL":nombre,"44578531":str(cfg["dni"]),
        "Estado civil: Soltera":f"Estado civil: {ec}","Soltera":ec,
        "Prolongación Arequipa 250, distrito y provincia de Barranca, departamento de Lima":cfg.get("domicilio",""),
        "Señora:":trato,"Estimada Sra. Caballero":estimado,
        "Lima, 11 de junio de 2026":f"Lima, {fecha_larga(d)}",
        "SEP-VP-203-2026-001":f"SEP-VP-{num}-{yyyy}-{cfg.get('n_sep','001')}",
        "PROP-VP-203-2026-01A-v2":f"PROP-VP-{num}-{yyyy}-01A","PROP-VP-203-2026-01B":f"PROP-VP-{num}-{yyyy}-01B",
        "(junio 2026)":f"({mes_anio(d)})",
        "Departamento Nº 203 – Piso 2 – Tipología 3D-A":f"Departamento Nº {num} – {piso} – Tipología {tip}",
        "Departamento N.° 203, Piso 2.":f"Departamento N.° {num}, {piso}.",
        "Departamento N° 203 (Unidad N° 05)":f"Departamento N° {num} (Unidad N° {un})",
        "Departamento Nº 203 (Unidad Nº 05)":f"Departamento Nº {num} (Unidad Nº {un})",
        "Piso 2 — Tipología 3D-A — 76 m²":f"{piso} — Tipología {tip} — {area} m²",
        "Piso 2 — Tipología 3D-A.":f"{piso} — Tipología {tip}.",
        "76 m² (con tolerancia de variación de hasta 3% por tratarse de bien futuro).":f"{area} m² (con tolerancia de variación de hasta 3% por tratarse de bien futuro).",
        "3 dormitorios (3D-A) — 76 m² de área techada.":f"{nd_txt} ({tip}) — {area} m² de área techada.",
        "tipología 3D-A,":f"tipología {tip},","área techada de 76 m²":f"área techada de {area} m²",
        "tres (03) dormitorios y dos (02) baños":f"{dorm} y {banos}",
        "en el piso 2 del proyecto VIVA PRO":f"en el {piso.lower()} del proyecto VIVA PRO",
        "Departamento 203":f"Departamento {num}",
        "320,000.00":m2(P),"128,000.00":m2(inicial40),"192,000.00":m2(saldo60),"160,000.00":m2(half50),
        "64,000.00":m2(p20),"60,500.00":m2(p20_sep),"48,000.00":m2(armada),"32,000.00":m2(p10),"3,500.00":m2(S),
        "S/ 64,000 menos separación":f"S/ {m0(p20)} menos separación",
        "1.09 %":pct(S/P*100),"18.91 %":pct(p20_sep/P*100),
        "TRESCIENTOS VEINTE MIL CON 00/100 SOLES":palabras_soles(P),
        "TRES MIL QUINIENTOS CON 00/100 SOLES":palabras_soles(S),"US$ 1,000.00":f"US$ {m2(sep_usd)}",
    }
    if cfg.get("conyuge"): R["No aplica – DNI No aplica"]=f"{cfg['conyuge']} – DNI {cfg.get('conyuge_dni','')}"
    if sexo=="M": R["la clienta"]="el cliente"
    return R

def generar_simulacion(cfg, dep, out, num, ape):
    tpl=os.path.join(TPL,"TPL_Simulador_Credito.xlsx")
    if not os.path.exists(tpl): return None
    tmpx=os.path.join(out,"_sim_tmp.xlsx")
    wb=openpyxl.load_workbook(tpl); ws=wb["Simulador"]
    ws["D5"]=cfg["nombre"].upper(); ws["D6"]=str(num)
    ws["D8"]=cfg.get("plazo_anios",20); ws["D9"]=cfg.get("tipo_cuota","Simple")
    ws["D12"]=cfg.get("tea_mivivienda",0.09); ws["D13"]=cfg.get("tea_tradicional",0.09)
    ws["D17"]=cfg.get("inicial_pct",0.20); ws["D22"]=float(cfg["precio_soles"])
    for h in ("Lista de Precios","Cronograma de Pago"):
        if h in wb.sheetnames: wb[h].sheet_state="hidden"
    ws.print_area="B1:E36"
    ws.page_setup.orientation="portrait"; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1
    ws.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
    wb.calculation.fullCalcOnLoad=True; wb.active=wb.sheetnames.index("Simulador")
    wb.save(tmpx); to_pdf(tmpx,out)
    final=os.path.join(out,f"Simulacion_Credito_Depa{num}_{ape}.pdf")
    if os.path.exists(os.path.join(out,"_sim_tmp.pdf")): os.replace(os.path.join(out,"_sim_tmp.pdf"),final)
    os.remove(tmpx); return os.path.basename(final)

def checklist(cfg, dep):
    P=cfg["precio_soles"]; bbp=dep.get("bbp")
    bbp_txt=f"S/ {bbp:,}" if bbp else "sin bono (dúplex / fuera de rango)"
    return f"""# Paquete de separación — Depto {dep['codigo']} ({dep['tipologia']}, {dep['area_m2']} m²)
**Cliente:** {cfg['nombre']} · DNI {cfg['dni']}
**Precio:** S/ {P:,.0f} · **Separación:** S/ {cfg.get('separacion_soles',3500):,.0f} · **BBP estimado:** {bbp_txt}

## A) Documentos que LA HAUS ENTREGA al cliente
- [ ] Propuesta comercial · [ ] Simulación de crédito · [ ] Ficha Técnica · [ ] Contrato de Separación (2 ejemplares) · [ ] Ficha PLAFT · [ ] Lista de precios

## B) Documentos que el CLIENTE ENTREGA
- [ ] DNI titular · [ ] DNI cónyuge (si aplica) · [ ] Ficha PLAFT firmada · [ ] Voucher separación (Interbank 200-3007447734) · [ ] Carta aprobación de crédito (si financia)

## C) Recordatorios
- Vigencia separación: 15 días calendario · Devolución íntegra solo por causa del Vendedor o denegatoria de crédito.
"""

def main():
    cfg=json.load(open(sys.argv[1],encoding="utf-8"))
    cat=json.load(open(os.path.join(TPL,"deptos.json"),encoding="utf-8"))
    cod=str(cfg["codigo_depto"]).strip()
    if cod not in cat: sys.exit(f"Código '{cod}' no está en catálogo. Disponibles: {', '.join(cat)}")
    dep=cat[cod]
    if not cfg.get("precio_soles"): cfg["precio_soles"]=dep["precio_final_soles"]
    out=cfg["carpeta_salida"]; os.makedirs(out,exist_ok=True)
    repl=construir(cfg,dep); ape=cfg.get("apellido","Cliente").replace(" ",""); num=dep["codigo"]
    planos_dir=cfg.get("planos_dir",DEFAULT_PLANOS)
    # swap del plano de la ficha
    ficha_swap=None
    if dep.get("plano"):
        pf=os.path.join(planos_dir,dep["plano"])
        if os.path.exists(pf):
            ficha_swap={"word/media/plano_distribucion.jpeg": fit_canvas(pf)}
    jobs=[]
    if "A" in cfg.get("opciones",["A","B"]): jobs.append(("TPL_Propuesta_OpcionA.docx",f"Propuesta_VIVA_PRO_Depa{num}_{ape}_OpcionA.docx",None))
    if "B" in cfg.get("opciones",["A","B"]): jobs.append(("TPL_Propuesta_OpcionB.docx",f"Propuesta_VIVA_PRO_Depa{num}_{ape}_OpcionB.docx",None))
    jobs.append(("TPL_Ficha_Tecnica.docx",f"Ficha_Tecnica_Depa{num}_VIVAPRO.docx",ficha_swap))
    jobs.append(("TPL_Contrato_Separacion.docx",f"Contrato_Separacion_{ape}_Depa{num}.docx",None))
    for tpl,name,swap in jobs:
        dx=os.path.join(out,name); render_docx(os.path.join(TPL,tpl),dx,repl,swap); to_pdf(dx,out); print("OK",name)
    if cfg.get("incluir_simulacion",True):
        s=generar_simulacion(cfg,dep,out,num,ape)
        if s: print("OK",s)
    open(os.path.join(out,"CHECKLIST_Reunion_Separacion.md"),"w",encoding="utf-8").write(checklist(cfg,dep))
    print("OK CHECKLIST"); print("Carpeta:",out)

if __name__=="__main__": main()
